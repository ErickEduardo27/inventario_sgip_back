"""Conversión CSV → Excel (.xlsx) con estilos y formatos de celda por columna."""

from __future__ import annotations

import io
import re
from datetime import date, datetime
from typing import Any, Literal

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.modules.inventory.csv_export import _sanitize_excel_value

ColumnFormat = Literal["integer", "date", "datetime", "currency_pen", "text"]

EXCEL_NUMBER_FORMATS: dict[ColumnFormat, str] = {
    "integer": "0",
    "date": "dd/mm/yyyy",
    "datetime": "dd/mm/yyyy hh:mm",
    "currency_pen": '"S/ "#,##0.00',
    "text": "@",
}

# Perfil: hoja de captura (cabeceras de hoja)
HOJA_CAPTURA_COLUMN_FORMATS: dict[str, ColumnFormat] = {
    "numero_hoja": "integer",
    "fecha": "date",
    "fecha_creacion": "datetime",
}

_DEFAULT_HEADER_COLOR = "2474F5"
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATA_ALIGN = Alignment(vertical="top", wrap_text=False)
_ZEBRA_FILL = PatternFill("solid", fgColor="F8FAFC")
_THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)

_DATE_PATTERNS = (
    "%Y-%m-%d",
    "%d/%m/%Y",
    "%d-%m-%Y",
    "%Y/%m/%d",
)
_DATETIME_PATTERNS = (
    "%d/%m/%Y %H:%M",
    "%d/%m/%Y %H:%M:%S",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d %H:%M",
)


def _normalize_header(name: str) -> str:
    return re.sub(r"\s+", "_", str(name).strip().lower())


def _format_display_header(name: str) -> str:
    """Primera letra mayúscula y espacios en lugar de guiones bajos."""
    label = str(name).strip().replace("_", " ")
    if not label:
        return label
    return label[0].upper() + label[1:]


def _header_fill(color_hex: str | None) -> PatternFill:
    raw = (color_hex or _DEFAULT_HEADER_COLOR).strip().lstrip("#").upper()
    if not re.fullmatch(r"[0-9A-F]{6}", raw):
        raw = _DEFAULT_HEADER_COLOR
    return PatternFill("solid", fgColor=raw)


def _resolve_column_format(header: str, profiles: dict[str, ColumnFormat]) -> ColumnFormat:
    key = _normalize_header(header)
    return profiles.get(key, "text")


def _parse_date(value: str) -> date | None:
    raw = value.strip()
    if not raw:
        return None
    date_part = raw.split(" ")[0] if " " in raw else raw
    for candidate in (raw, date_part):
        for fmt in _DATE_PATTERNS:
            try:
                return datetime.strptime(candidate, fmt).date()
            except ValueError:
                continue
    try:
        parsed = pd.to_datetime(raw, dayfirst=True, errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.date()
    except (ValueError, TypeError):
        return None


def _parse_datetime(value: str) -> datetime | None:
    raw = value.strip()
    if not raw:
        return None
    for fmt in _DATETIME_PATTERNS:
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    try:
        parsed = pd.to_datetime(raw, dayfirst=True, errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.to_pydatetime()
    except (ValueError, TypeError):
        return None


def _coerce_value(raw: object, fmt: ColumnFormat) -> tuple[Any, ColumnFormat]:
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None, fmt
    text = _sanitize_excel_value(str(raw).strip())
    if text == "":
        return None, fmt

    if fmt == "integer":
        try:
            return int(float(text.replace(",", ""))), "integer"
        except ValueError:
            return text, "text"

    if fmt == "currency_pen":
        cleaned = (
            text.replace("S/", "")
            .replace("s/", "")
            .replace(",", "")
            .strip()
        )
        try:
            return float(cleaned), "currency_pen"
        except ValueError:
            return text, "text"

    if fmt == "date":
        parsed = _parse_date(text)
        return (parsed if parsed is not None else text), ("date" if parsed is not None else "text")

    if fmt == "datetime":
        parsed = _parse_datetime(text)
        return (parsed if parsed is not None else text), ("datetime" if parsed is not None else "text")

    return text, "text"


def _auto_column_width(header: str, values: list[Any], fmt: ColumnFormat) -> float:
    max_len = len(header)
    for value in values[:500]:
        if value is None:
            continue
        if isinstance(value, datetime):
            display = value.strftime("%d/%m/%Y %H:%M")
        elif isinstance(value, date):
            display = value.strftime("%d/%m/%Y")
        elif isinstance(value, float):
            display = f"{value:,.2f}"
        else:
            display = str(value)
        max_len = max(max_len, len(display))
    cap = 48 if fmt == "text" else 22
    return min(max(max_len + 2, 10), cap)


def _populate_styled_sheet(
    ws,
    df: pd.DataFrame,
    profiles: dict[str, ColumnFormat],
    *,
    header_color_hex: str | None = None,
) -> None:
    headers = [str(h) for h in df.columns.tolist()]
    if not headers:
        return

    header_fill = _header_fill(header_color_hex)
    display_headers = [_format_display_header(header) for header in headers]

    for col_idx, display_header in enumerate(display_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=display_header)
        cell.fill = header_fill
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    parsed_columns: list[list[tuple[Any, ColumnFormat]]] = []
    for col_idx, header in enumerate(headers):
        fmt = _resolve_column_format(header, profiles)
        column_values: list[tuple[Any, ColumnFormat]] = []
        for raw in df.iloc[:, col_idx].tolist():
            column_values.append(_coerce_value(raw, fmt))
        parsed_columns.append(column_values)

    for row_idx in range(len(df)):
        excel_row = row_idx + 2
        for col_idx in range(len(headers)):
            value, fmt = parsed_columns[col_idx][row_idx]
            cell = ws.cell(row=excel_row, column=col_idx + 1)
            if value is None:
                cell.value = None
            elif fmt == "text":
                cell.value = str(value)
                cell.number_format = EXCEL_NUMBER_FORMATS["text"]
            else:
                cell.value = value
                cell.number_format = EXCEL_NUMBER_FORMATS[fmt]
            cell.alignment = _DATA_ALIGN
            cell.border = _THIN_BORDER
            if row_idx % 2 == 1:
                cell.fill = _ZEBRA_FILL

    for col_idx, display_header in enumerate(display_headers, start=1):
        values = [v for v, _ in parsed_columns[col_idx - 1]]
        sample_fmt = _resolve_column_format(headers[col_idx - 1], profiles)
        width = _auto_column_width(display_header, values, sample_fmt)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def csv_bytes_to_styled_xlsx_bytes(
    csv_payload: bytes,
    *,
    column_formats: dict[str, ColumnFormat] | None = None,
    sheet_title: str = "Exportación",
    header_color_hex: str | None = None,
) -> bytes:
    """Genera Excel con encabezado estilizado y formatos numéricos/fecha/moneda/texto."""
    profiles = column_formats or {}
    text = csv_payload.decode("utf-8-sig", errors="replace")
    df = pd.read_csv(io.StringIO(text), dtype=str, keep_default_na=False)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    _populate_styled_sheet(ws, df, profiles, header_color_hex=header_color_hex)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def csv_bytes_to_styled_xlsx_bytes_multi(
    sheets: list[tuple[bytes, str, dict[str, ColumnFormat] | None]],
    *,
    header_color_hex: str | None = None,
) -> bytes:
    """Genera un libro Excel con varias hojas estilizadas."""
    wb = Workbook()
    wb.remove(wb.active)
    for payload, title, column_formats in sheets:
        text = payload.decode("utf-8-sig", errors="replace")
        df = pd.read_csv(io.StringIO(text), dtype=str, keep_default_na=False)
        ws = wb.create_sheet(title=title[:31])
        _populate_styled_sheet(ws, df, column_formats or {}, header_color_hex=header_color_hex)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
