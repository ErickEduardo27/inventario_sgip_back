"""Excel del reporte APTOT por local con colores por SIT_PAT."""

from __future__ import annotations

import io
import re
from uuid import UUID

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from app.modules.inventory.excel_styled_export import (
    ColumnFormat,
    _normalize_header,
    _populate_styled_sheet,
)
from app.modules.inventory.reporte_aptot_local_status import APTOT_LOCAL_SIT_PAT_ROW_COLORS
from app.modules.tenants.theme import primary_hex_openpyxl

APTOT_LOCALES_COLUMN_FORMATS: dict[str, ColumnFormat] = {
    "hoja": "integer",
    "f.captura": "date",
    "fecha": "date",
    "depreciación": "currency_pen",
    "depreciacion": "currency_pen",
    "valor_neto": "currency_pen",
    "valor": "currency_pen",
}


def _sit_pat_row_fill(sit_pat: str) -> PatternFill | None:
    key = str(sit_pat or "").strip()
    color = APTOT_LOCAL_SIT_PAT_ROW_COLORS.get(key)
    if not color:
        return None
    raw = color.strip().lstrip("#").upper()
    if not re.fullmatch(r"[0-9A-F]{6}", raw):
        return None
    return PatternFill("solid", fgColor=raw)


def _apply_sit_pat_row_colors(ws, df: pd.DataFrame) -> None:
    headers = [str(h) for h in df.columns.tolist()]
    if not headers:
        return

    sit_pat_idx: int | None = None
    for idx, header in enumerate(headers):
        if _normalize_header(header) == "sit_pat":
            sit_pat_idx = idx
            break
    if sit_pat_idx is None:
        return

    sit_values = df.iloc[:, sit_pat_idx].tolist()
    for row_idx, raw in enumerate(sit_values):
        fill = _sit_pat_row_fill(str(raw or ""))
        if fill is None:
            continue
        excel_row = row_idx + 2
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=excel_row, column=col_idx).fill = fill


def csv_bytes_to_aptot_locales_xlsx_bytes(csv_payload: bytes, tenant_id: UUID) -> bytes:
    """Convierte CSV del reporte APTOT local a Excel con colores por SIT_PAT."""
    text = csv_payload.decode("utf-8-sig", errors="replace")
    df = pd.read_csv(io.StringIO(text), dtype=str, keep_default_na=False)

    wb = Workbook()
    ws = wb.active
    ws.title = "APTOT local"
    header_color = primary_hex_openpyxl(tenant_id)
    _populate_styled_sheet(
        ws,
        df,
        APTOT_LOCALES_COLUMN_FORMATS,
        header_color_hex=header_color,
        preserve_header_labels=True,
    )
    _apply_sit_pat_row_colors(ws, df)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
